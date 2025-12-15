import base64
from io import BytesIO
from openpyxl import load_workbook
from odoo import models, fields, api
from odoo.exceptions import UserError
import re


class WizardImportContact(models.TransientModel):
    _name = 'wizard.import.contact'
    _description = 'Wizard para importar contactos'

    file = fields.Binary("Archivo Excel", required=True)
    filename = fields.Char("Nombre del archivo")


    def action_import(self):
        self.ensure_one()

        if not self.file:
            raise UserError("Debe cargar un archivo Excel válido.")

        # 1- Decodificar archivo
        try:
            data = base64.b64decode(self.file)
        except Exception:
            raise UserError("No se pudo decodificar el archivo cargado.")

        # 2- Cargar archivo Excel con openpyxl
        try:
            workbook = load_workbook(BytesIO(data), data_only=True)
        except Exception:
            raise UserError("El archivo debe ser un Excel .xlsx válido.")

        sheet = workbook.active   # Primera hoja

        # 3-Recorrer filas (saltando la cabecera)
        for row in sheet.iter_rows(min_row=2, values_only=True):

            # Evitar filas vacías
            if not row or all(col is None for col in row):
                continue

            name_raw = row[0] or ""
            linea_raw = row[1] or ""
            description_raw = row[2] or ""

            if not name_raw:
                # Si no hay nombre, el contacto no se puede crear
                continue


            # Procesar tags (description)

            tag_ids = []
            description_raw_str = str(description_raw).strip()

            # Hace del texto con separación - en una lista
            tags_text = [t.strip() for t in re.split(r'\s*-\s*', description_raw_str) if t.strip()]


            for tag in tags_text:
                tag_rec = self.env['contact.tag'].search([('name', '=', tag)], limit=1)
                if not tag_rec:
                    tag_rec = self.env['contact.tag'].create({'name': tag})
                tag_ids.append(tag_rec.id)

            # linea (Many2one)

            linea_rec = False
            linea_raw_str = str(linea_raw).strip()

            if linea_raw_str:
                linea_rec = self.env['contact.linea'].search([('name', '=', linea_raw_str)], limit=1)
                if not linea_rec:
                    linea_rec = self.env['contact.linea'].create({'name': linea_raw_str})


            # Crear contacto

            self.env['res.partner'].create({
                'name': name_raw,
                'linea_id': linea_rec.id if linea_rec else False,
                'description_ids': [(6, 0, tag_ids)],
            })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': "Importación completada",
                'message': "Los contactos han sido importados correctamente.",
                'type': 'success',
                'sticky': False,
            }
        }
